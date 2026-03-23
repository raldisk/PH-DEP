"""
Export regional inequality data to a formatted Excel report.
Output: output/ph_inequality_report.xlsx

Usage:
    python scripts/export_excel.py
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DSN      = os.environ.get("DATABASE_URL", "postgresql://inequality:inequality@localhost:5432/ph_inequality")
OUT_DIR  = Path(__file__).parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "ph_inequality_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0D1B2A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F0F4F8")


def fetch(query: str) -> pd.DataFrame:
    with psycopg2.connect(DSN) as conn:
        return pd.read_sql(query, conn)


def style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx, _ in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(ws.max_row + 1, 100))
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))


def apply_conditional_formatting(ws, col_letter: str, start_row: int, end_row: int) -> None:
    """Green-White-Red scale for poverty incidence columns."""
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",   # green = low poverty
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",        # red = high poverty
        ),
    )


def main() -> None:
    print("Fetching data from PostgreSQL...")

    regional_poverty = fetch(open("sql/poverty_by_region.sql").read())
    income_deciles   = fetch(open("sql/fies_income_deciles.sql").read())
    grdp             = fetch("SELECT * FROM raw.grdp_regional ORDER BY year, region_code")
    sae_summary      = fetch("""
        SELECT region_name,
               COUNT(*) AS lgu_count,
               ROUND(AVG(poverty_incidence)::NUMERIC, 2) AS avg_poverty,
               ROUND(MIN(poverty_incidence)::NUMERIC, 2) AS min_poverty,
               ROUND(MAX(poverty_incidence)::NUMERIC, 2) AS max_poverty,
               SUM(CASE WHEN poverty_incidence <= 20 THEN 1 ELSE 0 END) AS lgus_below_20pct
        FROM raw.poverty_sae_municipal
        GROUP BY region_name
        ORDER BY avg_poverty DESC
    """)

    print(f"Writing to {OUT_PATH}...")
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        regional_poverty.to_excel(writer, sheet_name="Regional Poverty Pivot",  index=False)
        income_deciles.to_excel(writer,   sheet_name="Income Deciles by Region", index=False)
        grdp.to_excel(writer,             sheet_name="GRDP Regional",             index=False)
        sae_summary.to_excel(writer,      sheet_name="SAE Municipal Summary",     index=False)

    wb = load_workbook(OUT_PATH)

    for sheet_name, df in [
        ("Regional Poverty Pivot",   regional_poverty),
        ("Income Deciles by Region", income_deciles),
        ("GRDP Regional",            grdp),
        ("SAE Municipal Summary",    sae_summary),
    ]:
        ws = wb[sheet_name]
        style_sheet(ws, df)

    # Conditional formatting on poverty incidence columns
    ws_pov = wb["Regional Poverty Pivot"]
    for col_name, col_letter in [("pov_2021", "C"), ("pov_2023", "D")]:
        apply_conditional_formatting(ws_pov, col_letter, 2, ws_pov.max_row)

    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size // 1024
    print(f"Done → {OUT_PATH}  ({size_kb:,} KB)")
    print("Sheets: Regional Poverty Pivot · Income Deciles by Region · GRDP Regional · SAE Municipal Summary")


if __name__ == "__main__":
    main()
