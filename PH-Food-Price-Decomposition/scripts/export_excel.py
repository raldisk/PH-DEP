"""
Export food price data to formatted Excel report.
Output: output/ph_food_price_report.xlsx

Usage:
    python scripts/export_excel.py
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DSN      = os.environ.get("DATABASE_URL", "postgresql://food:food@localhost:5432/ph_food_prices")
OUT_DIR  = Path(__file__).parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "ph_food_price_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0D1B2A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F0F4F8")


def fetch(query: str) -> pd.DataFrame:
    with psycopg2.connect(DSN) as conn:
        return pd.read_sql(query, conn)


def style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(ws.max_row + 1, 50))
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))


def main() -> None:
    print("Fetching data from PostgreSQL...")

    monthly_prices = fetch(open("sql/price_trend_by_commodity.sql").read())
    seasonal       = fetch("""
        SELECT commodity_slug, month_num, month_name,
               ROUND(AVG(seasonal_index)::NUMERIC, 4) AS avg_seasonal_index
        FROM (
            SELECT *,
                   EXTRACT(MONTH FROM month)::INT AS month_num,
                   TO_CHAR(month::date, 'Mon')    AS month_name,
                   AVG(retail_price_php) / NULLIF(
                       AVG(AVG(retail_price_php)) OVER (PARTITION BY commodity_slug), 0
                   ) AS seasonal_index
            FROM (
                SELECT DATE_TRUNC('month',price_date)::DATE AS month,
                       commodity_slug, AVG(retail_price_php) AS retail_price_php
                FROM raw.psa_price_situationer
                WHERE region = 'National'
                GROUP BY 1, 2
            ) sub
        ) si
        GROUP BY commodity_slug, month_num, month_name
        ORDER BY commodity_slug, month_num
    """)
    shocks = fetch("""
        SELECT month, commodity_slug,
               ROUND(residual::NUMERIC, 4) AS residual,
               ROUND(z_score::NUMERIC, 3)  AS z_score
        FROM (
            SELECT month, commodity_slug, residual,
                   (residual - AVG(residual) OVER (PARTITION BY commodity_slug))
                   / NULLIF(STDDEV(residual) OVER (PARTITION BY commodity_slug), 0)
                   AS z_score
            FROM raw.stl_residuals
        ) z
        WHERE ABS(z_score) > 2
        ORDER BY ABS(z_score) DESC
        LIMIT 200
    """) if _table_exists("raw.stl_residuals", DSN) else pd.DataFrame()

    print(f"Writing to {OUT_PATH}...")
    sheets = [
        ("Monthly Price Trends", monthly_prices),
        ("Seasonal Index",       seasonal),
    ]
    if not shocks.empty:
        sheets.append(("Price Shocks (|z|>2σ)", shocks))

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(OUT_PATH)
    for sheet_name, df in sheets:
        style_sheet(wb[sheet_name], df)

    # Conditional formatting on seasonal index
    if "Seasonal Index" in wb.sheetnames:
        ws = wb["Seasonal Index"]
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="num", mid_value=1.0, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size // 1024
    print(f"Done → {OUT_PATH}  ({size_kb:,} KB)")
    for sheet_name, _ in sheets:
        print(f"  Sheet: {sheet_name}")


def _table_exists(table: str, dsn: str) -> bool:
    schema, tbl = table.split(".")
    try:
        with psycopg2.connect(dsn) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT 1 FROM information_schema.tables "
                    "WHERE table_schema=%s AND table_name=%s",
                    (schema, tbl)
                )
                return cur.fetchone() is not None
    except Exception:
        return False


if __name__ == "__main__":
    main()
