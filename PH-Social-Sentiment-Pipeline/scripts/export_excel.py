"""
Export sentiment mart data to a formatted Excel report.
Output: output/ph_sentiment_report.xlsx

Usage:
    python scripts/export_excel.py
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DSN      = os.environ.get("PH_SENTIMENT_POSTGRES_DSN",
                           "postgresql://sentiment:sentiment@localhost:5432/ph_sentiment")
OUT_DIR  = Path(__file__).parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "ph_sentiment_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="0D1B2A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F0F4F8")


def fetch(q: str) -> pd.DataFrame:
    with psycopg2.connect(DSN) as conn:
        return pd.read_sql(q, conn)


def style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(ws.max_row + 1, 50))
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))


def main() -> None:
    print("Fetching mart data...")

    trending   = fetch("SELECT * FROM marts.trending_topics ORDER BY day_manila DESC, daily_rank LIMIT 500")
    hourly_piv = fetch("""
        SELECT
            hour_manila,
            topic_name,
            positive, neutral, negative, total,
            ROUND(sentiment_score::NUMERIC, 4) AS sentiment_score
        FROM marts.sentiment_hourly
        ORDER BY hour_manila DESC, total DESC
        LIMIT 1000
    """)
    keywords   = fetch("SELECT * FROM marts.keyword_volume ORDER BY day_manila DESC, daily_rank LIMIT 200")
    raw_counts = fetch("""
        SELECT 'tweet_events' AS table_name, COUNT(*) AS row_count FROM raw.tweet_events
        UNION ALL
        SELECT 'trend_snapshots', COUNT(*) FROM raw.trend_snapshots
    """)

    print(f"Writing to {OUT_PATH}...")
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        trending.to_excel(writer,   sheet_name="Trending Topics",      index=False)
        hourly_piv.to_excel(writer, sheet_name="Hourly Sentiment",     index=False)
        keywords.to_excel(writer,   sheet_name="Keyword Volume",        index=False)
        raw_counts.to_excel(writer, sheet_name="Pipeline Row Counts",  index=False)

    wb = load_workbook(OUT_PATH)
    for sheet, df in [
        ("Trending Topics",     trending),
        ("Hourly Sentiment",    hourly_piv),
        ("Keyword Volume",      keywords),
        ("Pipeline Row Counts", raw_counts),
    ]:
        style_sheet(wb[sheet], df)
    wb.save(OUT_PATH)

    size_kb = OUT_PATH.stat().st_size // 1024
    print(f"Done → {OUT_PATH}  ({size_kb} KB)")


if __name__ == "__main__":
    main()
